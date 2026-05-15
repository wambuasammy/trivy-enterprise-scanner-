#!/usr/bin/env python3
"""
HardenX Excel Report Generator

Reads all CSV files in an assessment directory and creates a consolidated
Excel workbook with one worksheet per CSV file.
"""

import csv
import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def safe_sheet_name(name: str) -> str:
    """Convert filename to valid Excel worksheet name (max 31 chars)."""
    name = name.replace(".csv", "")
    name = name.replace("-", " ")
    return name[:31]


def auto_size_columns(ws):
    """Adjust column widths based on content."""
    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_length:
                max_length = len(value)

        ws.column_dimensions[column_letter].width = min(max_length + 2, 80)


def add_csv_to_sheet(workbook, csv_path: Path):
    """Create worksheet from CSV file."""
    ws = workbook.create_sheet(title=safe_sheet_name(csv_path.stem))

    with csv_path.open("r", newline="", encoding="utf-8", errors="replace") as f:
        reader = csv.reader(f)
        for row in reader:
            ws.append(row)

    auto_size_columns(ws)


def main():
    if len(sys.argv) != 2:
        print("Usage: excel_reporter.py <assessment_directory>")
        sys.exit(1)

    assessment_dir = Path(sys.argv[1])

    if not assessment_dir.is_dir():
        print(f"Directory not found: {assessment_dir}")
        sys.exit(1)

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    csv_files = sorted(assessment_dir.glob("*.csv"))

    if not csv_files:
        print("No CSV files found.")
        sys.exit(1)

    for csv_file in csv_files:
        add_csv_to_sheet(workbook, csv_file)

    output_file = assessment_dir / f"{assessment_dir.name}-assessment.xlsx"
    workbook.save(output_file)

    print(output_file)


if __name__ == "__main__":
    main()
